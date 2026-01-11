from flask import Flask, request, jsonify
from flask_cors import CORS
import os
from PIL import Image
import imagehash
import io
import re
import fitz  # PyMuPDF
from collections import defaultdict
import gc  # Garbage collector

app = Flask(__name__)
CORS(app)

# Configuración para optimizar memoria
MAX_IMAGE_SIZE = (800, 800)  # Reducir imágenes a máximo 800x800px
HASH_SIZE = 8  # Tamaño del hash perceptual

def resize_image(image, max_size=MAX_IMAGE_SIZE):
    """
    Redimensiona imagen manteniendo aspect ratio
    Reduce significativamente el uso de memoria
    """
    image.thumbnail(max_size, Image.Resampling.LANCZOS)
    return image

def extract_from_pdf(pdf_file):
    """
    Extrae productos de un catálogo PDF asociando imágenes con texto cercano
    """
    products = []
    pdf_document = None
    
    # Palabras clave que indican que NO es un producto
    HEADER_KEYWORDS = [
        'PRODUCTO', 'FOTO', 'MODELO', 'MAYOREO', 'MITAD', 'CAJA', 'CANTIDAD',
        'DESCRIPCION', 'DESCRIPCIÓN', 'PRECIO', 'MOQ', 'SKU', 'CODIGO',
        'CATEGORIA', 'DISPONIBILIDAD', 'PZAS', 'MENUDEO', 'LISTA', 'CODIGO DE PRODUCTO'
    ]
    
    try:
        # Leer el PDF en memoria
        pdf_bytes = pdf_file.read()
        pdf_document = fitz.open(stream=pdf_bytes, filetype="pdf")
        
        total_pages = len(pdf_document)
        print(f"📄 PDF con {total_pages} páginas")
        
        for page_num in range(total_pages):
            page = pdf_document[page_num]
            
            # Extraer imágenes con sus posiciones
            image_list = page.get_images(full=True)
            
            if not image_list:
                print(f"📄 Página {page_num + 1}: Sin imágenes, saltando...")
                continue
            
            # Extraer texto con posiciones (bloques)
            blocks = page.get_text("blocks")
            
            print(f"📄 Página {page_num + 1}/{total_pages}: {len(image_list)} imágenes")
            
            # Para cada imagen, buscar texto cercano
            for img_index, img in enumerate(image_list):
                try:
                    # Extraer la imagen
                    xref = img[0]
                    base_image = pdf_document.extract_image(xref)
                    image_bytes = base_image["image"]
                    
                    # Convertir a PIL Image
                    image = Image.open(io.BytesIO(image_bytes))
                    
                    # Optimizar imagen
                    image = resize_image(image)
                    if image.mode not in ('RGB', 'L'):
                        image = image.convert('RGB')
                    
                    # Calcular hash
                    img_hash = str(imagehash.phash(image, hash_size=HASH_SIZE))
                    
                    # Obtener posición de la imagen en la página
                    img_rect = page.get_image_bbox(img)
                    img_y = (img_rect.y0 + img_rect.y1) / 2  # Centro vertical de la imagen
                    
                    # Buscar texto en la misma fila (±50 puntos de diferencia)
                    row_texts = []
                    row_numbers = []
                    
                    for block in blocks:
                        if len(block) < 5:
                            continue
                        
                        x0, y0, x1, y1, text = block[0], block[1], block[2], block[3], block[4]
                        block_y = (y0 + y1) / 2
                        
                        # Si está en la misma fila que la imagen
                        if abs(block_y - img_y) < 50:
                            text_clean = text.strip()
                            
                            # Ignorar cabeceras
                            if any(keyword in text_clean.upper() for keyword in HEADER_KEYWORDS):
                                continue
                            
                            # Guardar texto
                            if len(text_clean) > 2:
                                row_texts.append(text_clean)
                            
                            # Buscar números (precios)
                            numbers = re.findall(r'\b(\d{2,4})\b', text_clean)
                            for num_str in numbers:
                                num = int(num_str)
                                if 10 <= num <= 5000:
                                    row_numbers.append(num)
                    
                    # Buscar descripción (texto más largo que no sea número)
                    description = None
                    for text in row_texts:
                        # Ignorar si es solo números
                        if text.replace(' ', '').replace('-', '').isdigit():
                            continue
                        # Ignorar si es muy corto
                        if len(text) < 5:
                            continue
                        # Tomar el texto más largo
                        if not description or len(text) > len(description):
                            description = text
                    
                    if not description:
                        description = f"Producto {img_index + 1}"
                    
                    # Buscar SKU (texto corto alfanumérico)
                    sku = None
                    for text in row_texts:
                        if 2 <= len(text) <= 10 and text.replace('-', '').replace('_', '').isalnum():
                            # No usar la descripción como SKU
                            if text != description:
                                sku = text
                                break
                    
                    if not sku:
                        sku = f"PROD-{page_num+1}-{img_index+1}"
                    
                    # Procesar precios
                    if len(row_numbers) >= 3:
                        prices_sorted = sorted(set(row_numbers))[:3]
                        price_mayoreo = prices_sorted[0]
                        price_mitad = prices_sorted[1]
                        price_caja = prices_sorted[2]
                    elif len(row_numbers) == 2:
                        prices_sorted = sorted(set(row_numbers))
                        price_mayoreo = prices_sorted[0]
                        price_mitad = prices_sorted[1]
                        price_caja = prices_sorted[1]
                    elif len(row_numbers) == 1:
                        price_mayoreo = row_numbers[0]
                        price_mitad = row_numbers[0]
                        price_caja = row_numbers[0]
                    else:
                        # Si no hay precios, saltar este producto
                        print(f"⚠️ Imagen {img_index+1}: Sin precios, saltando")
                        continue
                    
                    # Buscar MOQ
                    moq = 100
                    for text in row_texts:
                        if 'PIEZA' in text.upper() or 'DOCENA' in text.upper():
                            moq_match = re.search(r'(\d+)\s*(?:PIEZA|DOCENA|CAJITA)', text, re.IGNORECASE)
                            if moq_match:
                                moq = int(moq_match.group(1))
                                break
                    
                    # Detectar categoría
                    category = 'GENERAL'
                    desc_lower = description.lower()
                    if any(word in desc_lower for word in ['gorro', 'bufanda', 'caballero', 'dama', 'niño', 'niña', 'sombrero']):
                        category = 'ROPA Y ACCESORIOS'
                    elif any(word in desc_lower for word in ['navidad', 'luces', 'estrellitas', 'decoracion', 'regalo']):
                        category = 'DECORACION'
                    elif any(word in desc_lower for word in ['bolsa', 'empaque', 'papel']):
                        category = 'EMPAQUES Y REGALOS'
                    elif any(word in desc_lower for word in ['impermeable', 'tira', 'led', 'usb']):
                        category = 'ELECTRONICA'
                    
                    product = {
                        'sku': sku,
                        'description': description,
                        'priceMenudeo': round(float(price_mayoreo), 2),
                        'priceCaja': round(float(price_caja), 2),
                        'moq': moq,
                        'category': category,
                        'image_hash': img_hash,
                    }
                    
                    products.append(product)
                    print(f"✅ {description[:40]} | SKU: {sku} | Precios: {price_mayoreo}/{price_mitad}/{price_caja} | MOQ: {moq}")
                    
                    # Limpiar memoria
                    del image
                    del image_bytes
                    
                except Exception as e:
                    print(f"❌ Error en imagen {img_index+1}: {e}")
                    continue
            
            # Limpiar memoria
            del page
            if page_num % 5 == 0:
                gc.collect()
        
    except Exception as e:
        print(f"❌ Error leyendo PDF: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if pdf_document:
            pdf_document.close()
        gc.collect()
    
    print(f"📦 Total productos extraídos: {len(products)}")
    return products

def extract_from_excel(excel_file):
    """
    Procesar Excel de forma optimizada
    """
    from openpyxl import load_workbook
    
    products = []
    wb = None
    
    try:
        wb = load_workbook(excel_file, data_only=True)
        sheet = wb.active
        
        # Detectar columnas
        headers = [cell.value for cell in sheet[1]]
        
        col_mapping = {}
        for idx, header in enumerate(headers, 1):
            if header:
                h = str(header).lower()
                if 'sku' in h or 'codigo' in h or 'clave' in h:
                    col_mapping['sku'] = idx
                elif 'descripcion' in h or 'producto' in h or 'nombre' in h or 'description' in h:
                    col_mapping['description'] = idx
                elif 'menudeo' in h or 'retail' in h:
                    col_mapping['price'] = idx
                elif 'precio' in h and 'price' not in col_mapping:
                    col_mapping['price'] = idx
                elif 'caja' in h or 'mayoreo' in h or 'wholesale' in h:
                    col_mapping['priceCaja'] = idx
                elif 'moq' in h or 'minimo' in h or 'minimum' in h:
                    col_mapping['moq'] = idx
                elif 'categoria' in h or 'category' in h:
                    col_mapping['category'] = idx
        
        print(f"📊 Columnas detectadas: {col_mapping}")
        
        # Extraer imágenes
        for image in sheet._images:
            try:
                row = image.anchor._from.row + 1
                
                # Leer datos de la fila
                sku = sheet.cell(row, col_mapping.get('sku', 1)).value or f"XLS-{row}"
                description = sheet.cell(row, col_mapping.get('description', 2)).value or "Producto sin descripción"
                price = sheet.cell(row, col_mapping.get('price', 3)).value or 50.00
                price_caja = sheet.cell(row, col_mapping.get('priceCaja', 4)).value
                
                if not price_caja:
                    price_caja = float(price) * 0.85
                
                moq = sheet.cell(row, col_mapping.get('moq', 5)).value or 100
                category = sheet.cell(row, col_mapping.get('category', 6)).value or "GENERAL"
                
                # Convertir imagen
                img_bytes = image._data()
                img = Image.open(io.BytesIO(img_bytes))
                
                # **OPTIMIZACIÓN: Reducir tamaño**
                img = resize_image(img)
                
                if img.mode not in ('RGB', 'L'):
                    img = img.convert('RGB')
                
                img_hash = str(imagehash.phash(img, hash_size=HASH_SIZE))
                
                product = {
                    'sku': str(sku),
                    'description': str(description),
                    'priceMenudeo': float(price),
                    'priceCaja': float(price_caja),
                    'moq': int(moq),
                    'category': str(category),
                    'image_hash': img_hash,
                }
                
                products.append(product)
                
                # Limpiar
                del img
                del img_bytes
                
            except Exception as e:
                print(f"❌ Error en fila {row}: {e}")
                continue
        
    except Exception as e:
        print(f"❌ Error leyendo Excel: {e}")
    finally:
        if wb:
            wb.close()
        gc.collect()
    
    print(f"📦 Total productos del Excel: {len(products)}")
    return products

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'healthy', 'message': 'API optimizada - PDF y Excel'}), 200

@app.route('/api/consolidate', methods=['POST'])
def consolidate_catalogs():
    try:
        if 'files' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No se enviaron archivos'
            }), 400
        
        files = request.files.getlist('files')
        
        if not files:
            return jsonify({
                'success': False,
                'error': 'Lista de archivos vacía'
            }), 400
        
        print(f"📦 Recibidos {len(files)} archivos")
        
        all_products = []
        
        # **OPTIMIZACIÓN 6: Procesar archivo por archivo y limpiar memoria**
        for idx, file in enumerate(files):
            filename = file.filename.lower()
            
            # Extraer nombre del proveedor (eliminar _parte1, _parte2, etc)
            base_name = file.filename.split('.')[0]
            # Remover _parte, _part, -parte, -part seguido de números
            import re
            provider_name = re.sub(r'[_-]?(parte?|part)[_-]?\d+$', '', base_name, flags=re.IGNORECASE)
            
            print(f"📄 Procesando archivo {idx+1}/{len(files)}: {file.filename} (Proveedor: {provider_name})")
            
            try:
                if filename.endswith('.pdf'):
                    products = extract_from_pdf(file)
                elif filename.endswith(('.xlsx', '.xls')):
                    products = extract_from_excel(file)
                else:
                    print(f"⚠️ Formato no soportado: {filename}")
                    continue
                
                # Añadir nombre de proveedor
                for product in products:
                    product['provider'] = provider_name
                
                all_products.extend(products)
                print(f"✓ {len(products)} productos de {provider_name}")
                
                # **OPTIMIZACIÓN 7: Limpiar después de cada archivo**
                del products
                gc.collect()
                
            except Exception as e:
                print(f"❌ Error procesando {file.filename}: {e}")
                continue
        
        if not all_products:
            return jsonify({
                'success': False,
                'error': 'No se pudieron extraer productos de los archivos'
            }), 400
        
        print(f"📊 Total productos: {len(all_products)}")
        
        # Agrupar productos similares por hash de imagen
        groups = defaultdict(list)
        for product in all_products:
            groups[product['image_hash']].append(product)
        
        print(f"🔗 Grupos formados: {len(groups)}")
        
        # Consolidar: elegir el mejor precio de cada grupo
        consolidated = []
        for hash_key, group in groups.items():
            group_sorted = sorted(group, key=lambda x: x['priceCaja'])
            best = group_sorted[0]
            
            prices = [p['priceCaja'] for p in group]
            max_price = max(prices)
            savings = round(max_price - best['priceCaja'], 2)
            
            # Generar contenido optimizado para TikTok Shop
            description_clean = best['description'][:60].strip()
            
            # Título optimizado (máx 60 caracteres)
            optimized_title = f"{description_clean} | {best['category']}"[:60]
            
            # Descripción optimizada con emojis y formato TikTok
            optimized_description = f"""✨ {description_clean}

🎯 CARACTERÍSTICAS:
• Categoría: {best['category']}
• MOQ: {best['moq']} piezas
• Disponible con múltiples proveedores

💰 PRECIO:
• Menudeo: ${best['priceMenudeo']:.2f}
• Por Caja: ${best['priceCaja']:.2f}

📦 Envíos disponibles
✅ Calidad garantizada
🚀 Entrega rápida"""

            # Hashtags relevantes
            category_hashtags = {
                'ROPA Y ACCESORIOS': '#fashion #accesorios #moda #estilo',
                'DECORACION': '#decoracion #hogar #navidad #luces',
                'EMPAQUES Y REGALOS': '#regalo #empaque #bolsas #packaging',
                'ELECTRONICA': '#tech #electronica #gadgets #led',
                'GENERAL': '#productos #mayoreo #ventas'
            }
            
            base_hashtags = '#tiktokshop #mayoreo #preciosmayoreo #ventasonline'
            category_specific = category_hashtags.get(best['category'], '#productos')
            hashtags = f"{base_hashtags} {category_specific}"
            
            # Calcular margen sugerido (30-50%)
            suggested_retail_low = round(best['priceCaja'] * 1.3, 2)
            suggested_retail_high = round(best['priceCaja'] * 1.5, 2)
            
            consolidated_product = {
                'consolidated_sku': f"CONS-{str(len(consolidated) + 1).zfill(4)}",
                'description': best['description'],
                'priceMenudeo': best['priceMenudeo'],
                'priceCaja': best['priceCaja'],
                'moq': best['moq'],
                'category': best['category'],
                'provider': best['provider'],
                'num_providers': len(group),
                'savings': savings,
                'alternatives': [
                    {
                        'provider': p['provider'],
                        'sku': p['sku'],
                        'priceCaja': p['priceCaja']
                    }
                    for p in group
                ],
                # Campos optimizados para TikTok Shop
                'optimizedTitle': optimized_title,
                'optimizedDescription': optimized_description,
                'hashtags': hashtags,
                'suggestedRetailPrice': f"${suggested_retail_low:.2f} - ${suggested_retail_high:.2f}",
                'profitMargin': "30-50%"
            }
            
            consolidated.append(consolidated_product)
        
        # Limpiar grupos de memoria
        del groups
        del all_products
        gc.collect()
        
        # Ordenar por categoría
        consolidated.sort(key=lambda x: x['category'])
        
        # Calcular estadísticas
        stats = {
            'totalProducts': len(consolidated),
            'totalSavings': round(sum(p['savings'] for p in consolidated), 2),
            'duplicateProducts': len([p for p in consolidated if p['num_providers'] > 1]),
            'avgProviders': round(sum(p['num_providers'] for p in consolidated) / len(consolidated), 1) if consolidated else 0
        }
        
        print(f"✅ Consolidación completa: {stats['totalProducts']} productos únicos")
        
        return jsonify({
            'success': True,
            'consolidated': consolidated,
            'stats': stats
        })
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
    finally:
        # **OPTIMIZACIÓN 8: Siempre limpiar al final**
        gc.collect()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
